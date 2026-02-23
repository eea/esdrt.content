# -*- coding: utf-8 -*-
from collections import defaultdict
from itertools import cycle
from functools import partial
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from operator import attrgetter
from Products.Five.browser import BrowserView
from io import BytesIO
from zope.component import getUtility
from zope.schema.interfaces import IVocabularyFactory


XLS_SAMPLE_HEADER = (
    'Observation description', 'Country', 'CFR catedory code',
    'Inventory year', 'Gas', 'Review year', 'Fuel',
    'MS key category', 'EU key category',
    'Parameter', 'Description flags', 'Initial question text',
    'Author override',
)

DESC = 'Description of the observation'
CFR_CODE = '1A1'
INVENTORY_YEAR = '2018'
REVIEW_YEAR = '2018'
REFERENCE_YEAR = '2018'
QUESTION_TEXT = (
    'The text of an initial Q&A question. '
    'Leave empty if you do not wish to add an initial question.'
)


# UnicodeEncodeError
def decode(s):
    return s.decode('UTF-8', 'replace') if isinstance(s, str) else s


def _get_vocabulary(context, name):
    factory = getUtility(IVocabularyFactory, name=name)
    return factory(context)


def adjust_column_widths(worksheet):
    """Resize columns to match header text width."""
    widths: T.Dict[str, int] = defaultdict(int)

    for row in worksheet:
        for col, cell in enumerate(row, 1):
            if type(cell).__name__ == "MergedCell":
                continue  # ignore
            letter = get_column_letter(col)
            value = str(cell.value)
            width = len(value)
            if width > widths[letter]:
                widths[letter] = width

    for letter, width in widths.items():
        worksheet.column_dimensions[letter].width = width


def adjust_row_heights(worksheet):
    for idx, row in enumerate(worksheet, 1):
        for col, cell in enumerate(row, 1):
            if type(cell).__name__ == "MergedCell":
                continue  # ignore
            letter = get_column_letter(col)
            # calculate height based on text length compared to columnwidth
            width = int(worksheet.column_dimensions[letter].width)
            text_length = len(str(cell.value)) if cell.value is not None else 0
            lines_estimate = (round(text_length / width) + 1) + str(
                cell.value
            ).count("\n")
            calculated_height = (
                round(lines_estimate * cell.font.size) if lines_estimate else 0
            )
            # update height if needed (default is 12.75pt)
            current_height = worksheet.row_dimensions[idx].height or 20
            if current_height < calculated_height:
                worksheet.row_dimensions[idx].height = calculated_height


class GetSampleXLS(BrowserView):
    def populate_cells(self, sheet):
        get_vocabulary = partial(_get_vocabulary, self.context)
        get_title = attrgetter('title')

        header = XLS_SAMPLE_HEADER

        enable_key_category = self.context.enable_key_category

        if not enable_key_category:
            header = [v for v in header if "key category" not in v]

        fuel_voc = get_vocabulary('esdrt.content.fuel')
        # not a mandatory field, value can be none
        fuels = cycle(list(map(get_title, fuel_voc)) + [None])

        country_voc = get_vocabulary('esdrt.content.eea_member_states')
        gas_voc = get_vocabulary('esdrt.content.gas')
        parameter_voc = get_vocabulary('esdrt.content.parameter')
        description_flags_voc = get_vocabulary('esdrt.content.highlight')

        countries = list(map(get_title, country_voc))

        ms_key_categ = cycle(['True', None])
        eu_key_categ = cycle(['True', None])
        gas = '\n'.join(map(get_title, gas_voc))
        parameter = '\n'.join(map(get_title, parameter_voc))
        description_flags = cycle(
            ['\n'.join(map(get_title, description_flags_voc)), None]
        )

        sheet.append(header)
        for idx, country in enumerate(countries):
            # get a value based on the country index position
            ms_key_cat = next(ms_key_categ)
            eu_key_cat = next(eu_key_categ)
            desc_fl = next(description_flags)
            fuel = next(fuels)
            row_0 = [
                DESC,
                country,
                CFR_CODE,
                INVENTORY_YEAR, gas,
                REVIEW_YEAR,
                fuel,
            ]
            row_key_categ = [ms_key_cat, eu_key_cat]
            row_1 = [
                parameter,
                desc_fl,
                QUESTION_TEXT,
            ]

            if enable_key_category:
                row = row_0 + row_key_categ + row_1
            else:
                row = row_0 + row_1

            sheet.append(row)

    def __call__(self):
        wb = Workbook()
        sheet = wb.create_sheet('Observation', 0)

        self.populate_cells(sheet)

        adjust_column_widths(sheet)
        adjust_row_heights(sheet)

        xls = BytesIO()

        wb.save(xls)

        xls.seek(0)
        filename = 'observation_import_sample.xlsx'
        self.request.response.setHeader(
            'Content-type', 'application/vnd.ms-excel; charset=utf-8'
        )
        self.request.response.setHeader(
            'Content-Disposition', 'attachment; filename={0}'.format(filename)
        )
        return xls.read()
